import openpyxl
import os
import logging

class DataReader:
    @staticmethod
    def read_excel(file_path: str, sheet_name: str = "Sheet1"):
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        wb = openpyxl.load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in {file_path}")

        sheet = wb[sheet_name]
        data = []

        headers = [cell.value for cell in sheet[1]]
        if not headers:
            raise ValueError("No headers found in Excel sheet.")

        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = dict(zip(headers, row))
            data.append(row_data)

        logging.info(f"✅ Loaded {len(data)} rows from Excel: {file_path}")
        return data

